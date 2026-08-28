import os
import shutil
import pandas as pd
from flask import Flask, request, jsonify, render_template, send_from_directory
from openpyxl.styles import Border, Side
import traceback
from datetime import datetime
from urllib.parse import unquote
from concurrent.futures import ThreadPoolExecutor  # For parallel processing

# Initialize Flask app
app = Flask(__name__)
app.config['APPLICATION_NAME'] = "Daily Shipment Monitor Maker"
UPLOAD_FOLDER = "uploads"
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure the upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Function to clear uploads folder
def clear_uploads():
    try:
        for filename in os.listdir(UPLOAD_FOLDER):
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
    except Exception as e:
        print(f"Error clearing upload folder: {str(e)}")

@app.route('/')
def index():
    return render_template('index.html', title="Daily Shipment Monitor Maker")

@app.route('/upload', methods=['POST'])
def upload_file():
    clear_uploads()  # Clear the uploads folder before uploading the new file
    if 'file' not in request.files:
        return jsonify({"message": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"message": "No selected file"}), 400
    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(file_path)
    return jsonify({"message": "File uploaded successfully", "path": file_path}), 200

@app.route('/refine', methods=['POST'])
def refine_data():
    file_path = request.json.get('file_path')
    if not file_path or not os.path.exists(file_path):
        return jsonify({"message": "File not found"}), 400
    try:
        excel_data = pd.ExcelFile(file_path)
        refined_data = {}
        sheets_to_process = ["QAV RAWDATA", "OMV RAWDATA", "KSA RAWDATA", "UAE RAWDATA", "NBR RAWDATA"]

        # Process sheets in parallel
        with ThreadPoolExecutor() as executor:
            futures = []
            for sheet_name in sheets_to_process:
                if sheet_name in excel_data.sheet_names:
                    futures.append(executor.submit(process_sheet, excel_data, sheet_name))
            
            for future in futures:
                sheet_name, df = future.result()
                refined_data[sheet_name] = df

        # Write all refined data to the Excel file in one go
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
            for sheet_name, df in refined_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return jsonify({"message": "Data refinement completed successfully", "file_path": file_path}), 200
    except Exception as e:
        print(f"Error refining data: {str(e)}\n{traceback.format_exc()}")
        return jsonify({"message": "Error refining data", "error": str(e)}), 500

def process_sheet(excel_data, sheet_name):
    """Helper function to process a single sheet."""
    df = pd.read_excel(excel_data, sheet_name=sheet_name, dtype=str)
    df.fillna("EMPTY", inplace=True)
    df['COUNT'] = 1
    return sheet_name, df

@app.route('/download/<filename>')
def download_file(filename):
    try:
        # Decode the filename to handle URL encoding
        filename = unquote(filename)
        upload_folder = os.path.abspath(app.config['UPLOAD_FOLDER'])
        file_path = os.path.join(upload_folder, filename)

        # Check if the file exists and is readable
        if not os.path.exists(file_path):
            return jsonify({"message": "File not found", "error": "File not found in directory"}), 404
        if not os.access(file_path, os.R_OK):
            return jsonify({"message": "File cannot be read", "error": "Permission denied"}), 403

        return send_from_directory(upload_folder, filename, as_attachment=True)
    except Exception as e:
        print(f"Error during file download: {str(e)}")
        return jsonify({"message": "Error during file download", "error": str(e)}), 500

@app.route('/prepare', methods=['POST'])
def prepare_data():
    file_path = request.json.get('file_path')
    if not file_path or not os.path.exists(file_path):
        return jsonify({"message": "File not found"}), 400
    try:
        excel_data = pd.ExcelFile(file_path)
        today_date = datetime.today().strftime('%Y-%m-%d')
        output_filename = f"Daily Shipment monitor {today_date}.xlsx"
        output_path = os.path.join(UPLOAD_FOLDER, output_filename)

        with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
            for sheet_name in excel_data.sheet_names:
                df = pd.read_excel(excel_data, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)  # Save original data

                if sheet_name.endswith("RAWDATA"):
                    report_sheet_name = sheet_name.replace("RAWDATA", "REPORT")
                    if 'source_owner_name' in df.columns and any('status' in col.lower() for col in df.columns):
                        status_columns = [col for col in df.columns if 'status' in col.lower()]
                        status_column = status_columns[0] if status_columns else None

                        if status_column:
                            pivot_table = df.pivot_table(
                                index='source_owner_name',
                                columns=status_column,
                                values='COUNT',
                                aggfunc='sum',
                                fill_value=0
                            )

                            # Add totals row at the bottom
                            totals_row = pivot_table.sum().rename("Total")
                            pivot_table.loc["Total"] = totals_row

                            pivot_table.to_excel(writer, sheet_name=report_sheet_name)

                            worksheet = writer.sheets[report_sheet_name]

                            # Borders for pivot table
                            for row in worksheet.iter_rows():
                                for cell in row:
                                    cell.border = Border(
                                        left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'), 
                                        bottom=Side(style='thin')
                                    )

                            summary_start_row = worksheet.max_row + 3

                            # Calculate Total Awaiting by summing specific columns
                            total_awaiting = totals_row.get('AWAITING_DATA', 0) + totals_row.get('AWAITING_REVIEW', 0) + totals_row.get('AWAITING_PRODUCT_COUNT', 0)

                            # Adjust "Total Completed" based on sheet names
                            total_completed = totals_row.get('AWAITING_PROCESSING', 0) if sheet_name in ["KSA RAWDATA", "NBR RAWDATA", "OMV RAWDATA", "QAV RAWDATA"] else totals_row.get('COMPLETED_PROCESSING', 0)

                            # Calculating detailed summary
                            totals = {
                                "Total Completed": total_completed,
                                "Total Waiting": total_awaiting,
                                "Total Manually Parked": totals_row.get('MANUAL_INVESTIGATION_COMPLETE', 0),
                                "Total Failed": sum(totals_row.get(status, 0) for status in [
                                    'FAILED_INITIAL_PROCESSING', 'INVALID', 'FAILED_PRODUCT_COUNT',
                                    'FAILED_PRODUCT_COUNT_WITH_ERROR', 'FETCHED_WITH_ERROR',
                                    'FAILED_PROCESSING', 'MANUAL_MISSING_CODE_PAIRING'
                                ]),
                                "Total Unavailable": totals_row.get('UNAVAILABLE', 0),
                                "Grand Total": total_completed + total_awaiting + totals_row.get('MANUAL_INVESTIGATION_COMPLETE', 0) + sum(totals_row.get(status, 0) for status in [
                                    'FAILED_INITIAL_PROCESSING', 'INVALID', 'FAILED_PRODUCT_COUNT',
                                    'FAILED_PRODUCT_COUNT_WITH_ERROR', 'FETCHED_WITH_ERROR',
                                    'FAILED_PROCESSING', 'MANUAL_MISSING_CODE_PAIRING'
                                ]) + totals_row.get('UNAVAILABLE', 0)
                            }

                            summary_df = pd.DataFrame(list(totals.items()), columns=["Summary Status", "Summary Total"])
                            summary_df.to_excel(writer, sheet_name=report_sheet_name, startrow=summary_start_row, index=False)

                            # Borders for summary section
                            for row in worksheet.iter_rows(min_row=summary_start_row + 1, max_row=summary_start_row + len(summary_df) + 1, max_col=2):
                                for cell in row:
                                    cell.border = Border(
                                        left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'), 
                                        bottom=Side(style='thin')
                                    )

        return jsonify({"message": "Shipment report prepared successfully", "path": output_path, "filename": output_filename}), 200
    except Exception as e:
        print(f"Error preparing data: {str(e)}\n{traceback.format_exc()}")
        return jsonify({"message": "Error preparing data", "error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)